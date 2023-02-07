from openpyxl import Workbook
wb = Workbook()
sheet = wb.active
sheet.title = 'Hcl'
#j=0
#for i in range (10,60,10):
    #j+=1
    # sheet.cell(row=j,column=1).value=i  "prints in rows increment by 10"
    #sheet.cell(row = 1, column = j).value = i # column
#for row in sheet.iter_rows(min_row = 1,max_row=5,max_col =2,min_col=2):
#    for r in row:
#      r.value = 1
for col in sheet.iter_cols(min_col =1,max_col=5,max_row = 1,min_row=1):
    for r in col:
         for i in col:
           i=100
           r.value = 0
           r.value = r.value+100



wb.save("..//dataset/demoopenxlwrite.xlsx")