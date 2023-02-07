from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border,Side,Font,PatternFill,Alignment
workbook=load_workbook(filename="..//dataset//suprise.xlsx")
sheet=workbook.active
logo=Image("..//dataset//pic.png")
#logo1=Image("..//dataset//rose.png")
logo.height=250
logo.width=250
sheet.add_image(logo,"D10")
#logo1.height=350
#logo1.width=350
#sheet.add_image(logo1,"D20")
sheet.merge_cells('A1:D1')
cell = sheet['B2']
cell.value = "HAPPY ROSE-DAY MY ROSE"
thin = Side(border_style = "thin",color="000000")
double = Side(border_style = "double",color="ffffff")
cell.border=Border(top=thin,bottom=double)
cell.fill = PatternFill('solid',fgColor='DDDDDD')
cell.font = Font(b=True,color='FF0000')
cell.alignment =Alignment(horizontal = "center")
workbook.save(filename="..//dataset//suprise.xlsx")
