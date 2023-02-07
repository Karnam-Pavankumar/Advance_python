import openpyxl
from openpyxl import Workbook
import pandas as pd
from pandas import read_excel
from openpyxl.chart import LineChart,Reference
import numpy as np
wb=Workbook()
sheet1=wb.active
sheet1.title="JAN_Sales"
sh=wb.create_sheet(title = "FEB_Sales")
sheet3=wb.create_sheet(title = "merge_sales")
sales_data=[["days","profit"],
           [1,300000],
           [2,400000],
           [3,50000],
           [4,340000],
           [5,30000],
           [6,1000000],
           ]
for row in sales_data:
    sheet1.append(row)
sales_data2=[["days","profit"],
           [1,3000],
           [2,560000],
           [3,11110000],
           [4,3000],
           [5,30000],
           [6,4000],
           ]
for row in sales_data2:
    sh.append(row)
wb.save("..//dataset//hcl_sales.xlsx")








